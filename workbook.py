from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define headers (user-provided columns, with Entity and Guarantor added at start)
headers = [
    "Entity", "Guarantor",
    "Revenue", "Income", "Interest", "Depreciation", "Addbacks", "CapEx", "NOI",
    "Total Annual Debt Service", "CF After Debt Service", "DSCR",
    "Committed Debt", "Recourse", "Ownership",
    "Guarantor NOI", "Guarantor Annual Debt Service", "Guarantor CF After Debt Service",
    "Guarantor Debt Recourse", "Other Interest", "Amortization", "Distributions",
    "Total Pymt/Mo", "Guarantor Net Income", "Rate Shock DSCR", "",
    "Annual Rate Shock", "Notes",
    "Debt 1", "Am", "Type 1 ", " Rate 1", "Pymt/Mo 1", "Pymt/Mo Over 1", "Annual Debt 1",
    "Debt 2", "Am 2", "Type 2", "Rate 2", "Pymt/Month 2", "Pymt/Mo Override 2"
]
# Column letters for key metrics (after Entity A, Guarantor B)
# Revenue C, Income D, Interest E, Depreciation F, Addbacks G, CapEx H, NOI I, Total Annual Debt J, CF K, DSCR L, etc.
noi_col = 'I'
dscr_col = 'L'

# Add Setup sheet
setup = wb.create_sheet("Setup")
setup['A1'] = "Guarantor"
setup['B1'] = "Include?"
bold_font = Font(bold=True)
setup['A1'].font = bold_font
setup['B1'].font = bold_font
guarantors = ["Guarantor1", "Guarantor2", "Guarantor3", "Guarantor4", "Guarantor5"]
for row, g in enumerate(guarantors, start=2):
    setup[f'A{row}'] = g
    setup[f'B{row}'] = True if row <= 3 else False  # Example defaults

# Create raw data sheets for each guarantor
data_sheets = []
for i, g in enumerate(guarantors, start=1):
    sheet_name = f"{g}_Data"
    ws = wb.create_sheet(sheet_name)
    data_sheets.append(sheet_name)
    # Add headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
    # Auto-fill Guarantor column (B) with name from Setup
    setup_row = i + 1
    ws['B2'] = f"='Setup'!A{setup_row}"
    # Make it draggable (user can copy down)
    # Add data validation for Entity column (A) - will link to Master later

# Add Master_Entities sheet
master = wb.create_sheet("Master_Entities")
master['A1'] = "Entity"
master['B1'] = "Average NOI"
master['C1'] = "Average DSCR"
for col in 'ABC':
    master[f'{col}1'].font = bold_font

# Dynamic unique entities formula in A2 (spill down)
vstack_entities = 'VSTACK(' + ','.join([f'IF(Setup!$B${i+1}, {data_sheets[i]}!$A$2:$A$1048576, "")' for i in range(5)]) + ')'
vstack_filter = 'VSTACK(' + ','.join([f'IF(Setup!$B${i+1}, {data_sheets[i]}!$A$2:$A$1048576<>"", "")' for i in range(5)]) + ')'
master['A2'] = f'=UNIQUE(FILTER({vstack_entities}, {vstack_filter}))'

# Average NOI in B2
vstack_noi = 'VSTACK(' + ','.join([f'IF(Setup!$B${i+1}, {data_sheets[i]}!{noi_col}:{noi_col}, "")' for i in range(5)]) + ')'
master['B2'] = f'=IF(A2="","", AVERAGEIFS({vstack_noi}, {vstack_entities}, A2))'

# Average DSCR in C2
vstack_dscr = 'VSTACK(' + ','.join([f'IF(Setup!$B${i+1}, {data_sheets[i]}!{dscr_col}:{dscr_col}, "")' for i in range(5)]) + ')'
master['C2'] = f'=IF(A2="","", AVERAGEIFS({vstack_dscr}, {vstack_entities}, A2))'

# Add Dashboard sheet
dashboard = wb.create_sheet("Dashboard")
dashboard['A1'] = "Select Guarantor 1"
dashboard['B1'] = "Select Guarantor 2"
dashboard['A1'].font = bold_font
dashboard['B1'].font = bold_font

# Dropdowns for guarantors (data validation from Setup)
dv = DataValidation(type="list", formula1='=Setup!$A$2:$A$6')
dashboard.add_data_validation(dv)
dv.add('A2')
dv.add('B2')

# Comparison table starting at A4
dashboard['A4'] = "Entity"
dashboard['B4'] = "G1 NOI"
dashboard['C4'] = "G1 DSCR"
dashboard['D4'] = "G2 NOI"
dashboard['E4'] = "G2 DSCR"
dashboard['F4'] = "NOI Variance"
dashboard['G4'] = "DSCR Variance"
for col in range(1, 8):
    dashboard.cell(row=4, column=col).font = bold_font

# Formulas for variances
dashboard['F5'] = '=IF(OR(B5="N/A",D5="N/A"),"N/A",ABS(B5-D5))'
dashboard['G5'] = '=IF(OR(C5="N/A",E5="N/A"),"N/A",ABS(C5-E5))'

# Copy all comparison formulas (B5:G5) down for 500 rows
formula_cells = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5']
for source_cell in formula_cells:
    formula = dashboard[source_cell].value
    if formula:
        for row in range(5, 505):
            dashboard.cell(row=row, column=dashboard[source_cell].column, value=formula)

# Optional: Improve entity spill reference (use full spill array for better auto-expansion)
dashboard['A5'] = '=Master_Entities!A2#'

# Freeze panes and set column widths for better usability
dashboard.freeze_panes = 'A5'
for col in range(1, 8):
    dashboard.column_dimensions[get_column_letter(col)].width = 20

# Add instructions in the Dashboard for users
dashboard['A1'] = "Guarantor Comparison Dashboard"
dashboard['A1'].font = Font(bold=True, size=14)
dashboard['I1'] = "Instructions:"
dashboard['I2'] = "1. Select two guarantors from dropdowns in A2 and B2."
dashboard['I3'] = "2. Paste raw data into the respective GuarantorX_Data sheets (Entity in column A)."
dashboard['I4'] = "3. Variances > 0.01 will be highlighted after you add conditional formatting (see below)."
dashboard['I6'] = "Add Conditional Formatting:"
dashboard['I7'] = "- Select F5:F504 > Home > Conditional Formatting > Greater Than > 0.01 > Red Fill"
dashboard['I8'] = "- Repeat for G5:G504"

# Save the template
wb.save("credit_underwriting_template.xlsx")
print("Template generated successfully!")
print("Open credit_underwriting_template.xlsx in Excel.")
print("- Dropdowns in Dashboard A2/B2 for selecting guarantors.")
print("- Add the suggested conditional formatting for variances.")
print("- Formulas will auto-update as you paste data.")